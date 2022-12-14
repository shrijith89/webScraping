import openpyxl
import requests
from bs4 import BeautifulSoup

sourceCode = requests.get("https://www.imdb.com/chart/top")
parsedData = BeautifulSoup(sourceCode.text, 'html.parser')
movieDetails = parsedData.find('tbody', class_="lister-list").find_all('tr')
excel = openpyxl.Workbook()
sheet = excel.active
sheet.title = "Top Charts Movies"
sheet.append(['Movie Rank', 'Movie Name/Year', 'Movie Ratings'])

for details in movieDetails:
    movieName = details.find('td', class_='titleColumn').find('a').text
    movieRank = details.find('td', class_='titleColumn').get_text(strip=True).split('.')[0]
    movieRelease = details.findNext('td', class_='titleColumn').find('span').text
    movieRating = details.findNext('td', class_='ratingColumn imdbRating').find('strong').text
    movieNameRelease = movieName + '-' + movieRelease
    sheet.append([movieRank, movieNameRelease, movieRating])

excel.save('IMDBRatings.xlsx')

w = openpyxl.load_workbook("IMDBRatings.xlsx")
worksheet = w.active


def searchmovies(searchMovie):
    for i in range(0, worksheet.max_row):
        for j in worksheet.iter_cols(1, worksheet.max_column - 1):
            if j[i].value == searchMovie:
                return j[i].value + ", The rank is " + worksheet.cell(row=i + 1, column=1).value
    return 0

print(searchmovies("Fight Club-(1999)"))
